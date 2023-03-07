# Change the values here as necessary
NAME_ISSUER = "Lee Xin Yi"
CLUB_CODE = "SAC110"
INPUT_FILE_NAME = "input.xlsx"
FONT_NAME = "Helvetica-Bold"
FONT_SIZE = 13
CALCULATE_GST = True
DATE_FORMAT = "%d/%m/%Y"

# Code
import io
import os
import inflect
import fitz
import openpyxl
from datetime import datetime
from docx import Document
from docx.shared import Cm
from reportlab.pdfgen import canvas
from PyPDF2 import PdfFileWriter, PdfFileReader

pdf_filenames = []
png_filenames = []
p = inflect.engine()

def main():
    wb_obj = openpyxl.load_workbook(INPUT_FILE_NAME)
    if not os.path.exists("out-pdf"):
        os.makedirs("out-pdf")
    if not os.path.exists("out-png"):
        os.makedirs("out-png")
    sheet_obj = wb_obj.active
    for row_num in range(2, sheet_obj.max_row + 1):
        receipt_num = sheet_obj.cell(row=row_num, column=1).value
        receipt_from = sheet_obj.cell(row=row_num, column=2).value
        receipt_amount = sheet_obj.cell(row=row_num, column=3).value
        receipt_description = sheet_obj.cell(row=row_num, column=4).value
        receipt_date = sheet_obj.cell(row=row_num, column=5).value
        generate_receipt_pdf(
            frm=receipt_from,
            amount=receipt_amount,
            receipt_no=receipt_num,
            description=receipt_description,
            date=receipt_date
        )
    convert_to_png(pdf_filenames)
    combine_into_doc(png_filenames)


def generate_receipt_pdf(frm:str="Payer Name", amount:float=0.0, receipt_no:int=12345, description:str = "ITEM NAME HERE",date:str=datetime.now().strftime(DATE_FORMAT)):
    print("Processing receipt number", receipt_no, "-", frm)

    # Create overlay PDF
    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=(1900,570))
    can.setFillColorRGB(0,0,0)
    can.setFont(FONT_NAME, FONT_SIZE)

    # Text Processing
    intpart,decimalpart = int(amount), amount-int(amount)
    dollars = p.number_to_words(intpart) + (" dollars" if intpart > 1 else " dollar")
    cents = " and " + p.number_to_words(int(decimalpart * 100)) + " cents"
    if decimalpart != 0:
        amount_text = (dollars + cents + " only").upper()
    else:
        amount_text = (dollars + " only").upper()
    included_gst = round((amount / 107) * 7, 2)

    if type(date) == datetime:
        date = date.strftime(DATE_FORMAT)

    # Draw on overlay PDF
    can.drawString(150, 145, f"{frm.title()}")
    can.drawString(150, 120, f"{amount_text.title()}")
    can.drawString(150, 95, f"{description}")
    can.drawString(50, 65, f"{amount:.2f}")
    can.drawString(675, 44, f"{NAME_ISSUER.title()}")
    can.drawString(675, 165, f"{date.title()}")
    can.drawString(675, 200, f"{CLUB_CODE.upper()} - {str(receipt_no).zfill(2)}")

    if CALCULATE_GST:
        can.drawString(122, 22, f"{included_gst:.2f}")

    can.save()

    # move to the beginning of the StringIO buffer
    packet.seek(0)
    new_pdf = PdfFileReader(packet)

    # read your existing PDF
    existing_pdf = PdfFileReader(open("smusa-base.pdf", "rb"))

    # add the "watermark" (which is the new pdf) on the existing page
    page = existing_pdf.getPage(0)
    page2 = new_pdf.getPage(0)
    page.mergePage(page2)

    # finally, write "output" to a real file
    output = PdfFileWriter()
    output.addPage(page)
    with open("out-pdf/" + f"RECEIPT - {receipt_no} - {frm}.pdf".replace("/", ""), "wb") as pdfFile:
        output.write(pdfFile)
        pdfFile.close()
        pdf_filenames.append(f"RECEIPT - {receipt_no} - {frm}.pdf".replace("/", ""))


def convert_to_png(pdf_filenames):
    for source in pdf_filenames:
        doc = fitz.open("out-pdf/" + source)
        page = doc.load_page(0)
        pix = page.get_pixmap()
        output = f"out-png/{source}.png"
        png_filenames.append(output)
        pix.save(output)


def combine_into_doc(png_filenames):
    document = Document()
    sections = document.sections
    for section in sections:
        section.top_margin = Cm(1.27)
        section.bottom_margin = Cm(1.27)
        section.left_margin = Cm(1.27)
        section.right_margin = Cm(1.27)
    for source in png_filenames:
        document.add_picture(source, width=Cm(19))
    try:
        document.save("output.docx")
    except PermissionError:
        print("ERROR: Cant save the docx file, is the file open?")


if __name__ == "__main__":
    main()